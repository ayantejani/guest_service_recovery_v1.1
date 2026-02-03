"""
Excel file parser for complaint data.
Equivalent to server/reportRouter.ts parseExcel procedure.
"""

from typing import List, Tuple
import openpyxl
from openpyxl.utils import get_column_letter
from app.report_utils import ParsedComplaint, parse_excel_row


def parse_excel_file(file_path: str) -> Tuple[List[ParsedComplaint], List[str]]:
    """
    Parse Excel file and extract complaint records.
    Handles HIEX format where headers are in row 3.
    
    Returns:
        Tuple of (complaints list, errors list)
    """
    complaints: List[ParsedComplaint] = []
    errors: List[str] = []

    try:
        # Load workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        if not sheet:
            errors.append("No active sheet found in Excel file")
            return complaints, errors

        # Get headers from row 3 (HIEX format)
        headers = []
        for cell in sheet[3]:
            headers.append(cell.value)

        # Parse data rows (start from row 4, skip header rows)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=4, values_only=False), start=4):
            try:
                # Convert row to dictionary
                row_dict = {}
                for col_idx, cell in enumerate(row, start=1):
                    header = headers[col_idx - 1] if col_idx - 1 < len(headers) else None
                    if header:
                        row_dict[header] = cell.value

                # Parse the row
                parsed = parse_excel_row(row_dict)
                if parsed:
                    complaints.append(parsed)
                else:
                    # Skip rows that don't parse (empty rows, etc.)
                    pass
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        if not complaints:
            errors.append("No valid complaints found in Excel file")

        return complaints, errors

    except Exception as e:
        errors.append(f"Failed to parse Excel file: {str(e)}")
        return complaints, errors


def parse_excel_from_bytes(file_bytes: bytes) -> Tuple[List[ParsedComplaint], List[str]]:
    """
    Parse Excel file from bytes.
    
    Returns:
        Tuple of (complaints list, errors list)
    """
    import tempfile
    import os

    complaints: List[ParsedComplaint] = []
    errors: List[str] = []

    try:
        # Write bytes to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name

        try:
            complaints, errors = parse_excel_file(tmp_path)
        finally:
            # Clean up temporary file
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

        return complaints, errors

    except Exception as e:
        errors.append(f"Failed to parse Excel from bytes: {str(e)}")
        return complaints, errors
